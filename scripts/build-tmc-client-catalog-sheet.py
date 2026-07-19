#!/usr/bin/env python3
"""Build client-facing Excel + CSV from TMC ring catalog scrape.

Columns for client editing:
  INCLUDE | NAME | PRICE | NOTES | (+ reference columns)

Usage: python3 scripts/build-tmc-client-catalog-sheet.py
"""

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
except ImportError:
    print("Install openpyxl: pip3 install openpyxl")
    raise

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "exports/tmc-ring-catalog"
CATALOG_PATH = OUT_DIR / "catalog.json"
XLSX_PATH = OUT_DIR / "TMC_Ring_Catalog_Client.xlsx"
CSV_PATH = OUT_DIR / "TMC_Ring_Catalog_Client.csv"

COLUMNS = [
    ("include", "INCLUDE (Yes/No)"),
    ("name", "NAME"),
    ("price", "PRICE"),
    ("notes", "NOTES"),
    ("category", "Category"),
    ("metalColour", "Metal Colour"),
    ("tmcOriginalName", "TMC Original Name"),
    ("tmcPriceAud", "TMC Price (AUD ref)"),
    ("imageFile", "Image File"),
    ("imageRelative", "Image Path"),
    ("handle", "TMC Handle"),
]


def main():
    if not CATALOG_PATH.exists():
        raise SystemExit(f"Missing {CATALOG_PATH}. Run: node scripts/scrape-tmc-ring-catalog.mjs")

    catalog = json.loads(CATALOG_PATH.read_text())
    rows = catalog.get("rows", [])

    # CSV (Google Sheets import)
    header = ",".join(f'"{label}"' for _, label in COLUMNS)
    lines = [header]
    for row in rows:
        vals = []
        for key, _ in COLUMNS:
            v = str(row.get(key, "")).replace('"', '""')
            vals.append(f'"{v}"')
        lines.append(",".join(vals))
    CSV_PATH.write_text("\ufeff" + "\n".join(lines) + "\n", encoding="utf-8")

    # Excel with hyperlinks to local images
    wb = Workbook()
    ws = wb.active
    ws.title = "Ring Catalog"

    header_fill = PatternFill(start_color="6D3D0D", end_color="6D3D0D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (_, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (key, _) in enumerate(COLUMNS, 1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Hyperlink image path column to local file
            if key == "imageRelative" and value:
                img_path = OUT_DIR / value
                if img_path.exists():
                    cell.hyperlink = str(img_path.resolve())
                    cell.font = Font(color="0563C1", underline="single")

    for col_idx, (_, label) in enumerate(COLUMNS, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(50, max(12, len(label) + 2))

    ws.freeze_panes = "A2"

    with_images = sum(1 for r in rows if r.get("imageRelative"))

    # Instructions sheet for client / Drive upload
    guide = wb.create_sheet("How to use")
    guide["A1"] = "TMC Ring Catalog — Client Instructions"
    guide["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "1. Upload the entire 'tmc-ring-catalog' folder (or zip) to your shared Google Drive.",
        "2. Open TMC_Ring_Catalog_Client.csv in Google Sheets (File → Import).",
        "3. Fill in columns: INCLUDE (Yes/No), NAME, PRICE, NOTES.",
        "4. One row = one ring in one metal colour (Yellow / White / Rose). Delete rows you don't need.",
        "5. Image Path column matches files inside the images/ folder on Drive.",
        "6. TMC Price (AUD ref) is the original setting price from TMC — use as a guide only.",
        "",
        f"Total rings scraped: {len({r['handle'] for r in rows})}",
        f"Total rows (ring × metal): {len(rows)}",
        f"Rows with images: {with_images}",
        "",
        "Lifestyle / on-hand photos were excluded. Only 3D metal renders are included.",
    ]
    for i, line in enumerate(lines, 2):
        guide.cell(row=i, column=1, value=line)
    guide.column_dimensions["A"].width = 90

    wb.save(XLSX_PATH)

    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote {CSV_PATH}")
    print(f"  rows: {len(rows)} ({with_images} with images)")
    print(f"\nUpload folder to Google Drive:")
    print(f"  {OUT_DIR}")
    print("Then import CSV into Google Sheets, or open XLSX in Google Sheets.")


if __name__ == "__main__":
    main()
