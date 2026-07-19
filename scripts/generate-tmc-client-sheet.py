#!/usr/bin/env python3
"""Write TMC client review sheet as .xlsx (requires openpyxl)."""
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
PRODUCTS_PATH = ROOT / "public/data/products.json"
MANIFEST_PATH = ROOT / "public/data/tmc-import-manifest.json"
OUT_PATH = ROOT / "docs/TMC_Ring_Import_Client_Sheet.xlsx"

COLUMNS = [
    ("includeOnSite", "Include on ELYSIUM site? (Yes/No)"),
    ("tmcOriginalName", "TMC Original Name"),
    ("tmcShopifyHandle", "TMC Shopify Handle"),
    ("elysiumTitle", "ELYSIUM Site Name (Title)"),
    ("elysiumSlug", "ELYSIUM URL Slug"),
    ("elysiumUrl", "ELYSIUM Product URL (when live)"),
    ("blurb", "Short Tagline (Blurb)"),
    ("description", "Full Product Description"),
    ("basePriceGBP", "Base Price GBP (from, 1ct lab-grown)"),
    ("featured", "Featured on Homepage? (Yes/No)"),
    ("shape", "Stone Shape"),
    ("styles", "Style Tags (comma-separated)"),
    ("collections", "Collections (comma-separated)"),
    ("metals", "Available Metals (comma-separated)"),
    ("carats", "Carat Options (comma-separated)"),
    ("origins", "Stone Origin Options (Natural / Lab Grown)"),
    ("colours", "Diamond Colour Grades Offered"),
    ("clarities", "Diamond Clarity Grades Offered"),
    ("certificates", "Certification Options (GIA / IGI)"),
    ("engraving", "Engraving Available? (Yes/No)"),
    ("qualityBanner", "Quality Banner Text"),
    ("seoTitle", "SEO Page Title"),
    ("seoDescription", "SEO Meta Description"),
    ("hasYellowImages", "Has Yellow Gold Images? (Yes/No)"),
    ("hasRoseImages", "Has Rose Gold Images? (Yes/No)"),
    ("hasWhitePlatinumImages", "Has White Gold / Platinum Images? (Yes/No)"),
    ("needsPhotography", "Needs New Photography? (Yes/No)"),
    ("clientNotes", "Client Notes / Special Instructions"),
]


def has_metal(product, key):
    return bool(product.get("galleryByMetal", {}).get(key))


def build_rows(products, manifest):
    handle_to_product = {}
    for p in products:
        handle = None
        for img in p.get("images", []):
            if "tmc-import/" in img:
                handle = img.split("tmc-import/")[1].split("/")[0]
                break
        if handle:
            handle_to_product[handle] = p

    site_url = "https://elysium-mvp.vercel.app"
    rows = []
    for entry in manifest["products"]:
        p = handle_to_product.get(entry["handle"])
        if not p:
            continue
        has_yellow = has_metal(p, "18k Yellow Gold")
        has_rose = has_metal(p, "18k Rose Gold")
        has_white = has_metal(p, "18k White Gold") or has_metal(p, "Platinum")
        rows.append({
            "includeOnSite": "",
            "tmcOriginalName": entry["title"],
            "tmcShopifyHandle": entry["handle"],
            "elysiumTitle": p["title"],
            "elysiumSlug": p["slug"],
            "elysiumUrl": f"{site_url}/products/{p['slug']}",
            "blurb": p.get("blurb", ""),
            "description": p.get("description", "").replace("\n", " ").strip(),
            "basePriceGBP": p.get("basePriceGBP", ""),
            "featured": "Yes" if p.get("isFeatured") else "No",
            "shape": p.get("shape", ""),
            "styles": ", ".join(p.get("styles", [])),
            "collections": ", ".join(p.get("collections", [])),
            "metals": ", ".join(m["name"] for m in p.get("metals", [])),
            "carats": ", ".join(c["label"] for c in p.get("carats", [])),
            "origins": ", ".join(o["label"] for o in p.get("origins", [])),
            "colours": ", ".join(c["label"] for c in p.get("colours", [])),
            "clarities": ", ".join(c["label"] for c in p.get("clarities", [])),
            "certificates": ", ".join(c["label"] for c in p.get("certificates", [])),
            "engraving": "Yes" if p.get("engravingMaxChars") else "No",
            "qualityBanner": p.get("qualityBanner", ""),
            "seoTitle": p.get("seoTitle", ""),
            "seoDescription": p.get("seoDescription", ""),
            "hasYellowImages": "Yes" if has_yellow else "No",
            "hasRoseImages": "Yes" if has_rose else "No",
            "hasWhitePlatinumImages": "Yes" if has_white else "No",
            "needsPhotography": "No" if (has_yellow and has_rose and has_white) else "Yes",
            "clientNotes": "",
        })
    return rows


def main():
    products = json.loads(PRODUCTS_PATH.read_text())
    manifest = json.loads(MANIFEST_PATH.read_text())
    rows = build_rows(products, manifest)

    wb = Workbook()
    ws = wb.active
    ws.title = "TMC Ring Import"

    header_fill = PatternFill(start_color="6D3D0D", end_color="6D3D0D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (_, header) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (key, _) in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    for col_idx, (_, header) in enumerate(COLUMNS, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(48, max(14, len(header) + 2))

    ws.freeze_panes = "A2"
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({len(rows)} rings)")


if __name__ == "__main__":
    main()
